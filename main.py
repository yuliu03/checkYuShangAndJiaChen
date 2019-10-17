from openpyxl import load_workbook
from openpyxl import Workbook

#input1Path 入库单明细表
#input2Path 进销存表


def myfilter(x,input2_sheet):
    if "=" not in str(x):
        # print("ok: "+str(x))
        return str(x)
    else:
        # print("not: "+str(x))
        if 'B' in x:
            return myfilter(input2_sheet.cell(row=int(str(x).replace('=', '').replace('B','')), column=2).value,input2_sheet)
        elif 'A' in x:
            return myfilter(input2_sheet.cell(row=int(str(x).replace('=', '').replace('A','')), column=1).value,input2_sheet)

input1Path = "C:/Users/admin/Desktop/input.xlsx"

input1_wb = load_workbook(input1Path)

input1_sheet = input1_wb[input1_wb.sheetnames[0]]
input2_sheet = input1_wb[input1_wb.sheetnames[1]]

rows = input1_sheet.max_row
pack1 = dict()
tmpRow = 2


while tmpRow <= rows:
    if input1_sheet.cell(row=tmpRow, column=2).value != None:
        if input1_sheet.cell(row=tmpRow, column=2).value not in pack1.keys():
            pack1[input1_sheet.cell(row=tmpRow, column=2).value] = (
                input1_sheet.cell(row=tmpRow, column=1).value,
                int(input1_sheet.cell(row=tmpRow, column=3).value)
            )
        else:
            pack1[input1_sheet.cell(row=tmpRow, column=2).value] = (
                input1_sheet.cell(row=tmpRow, column=1).value,
                pack1[input1_sheet.cell(row=tmpRow, column=2).value][1] +
                int(input1_sheet.cell(row=tmpRow, column=3).value)
            )
    else:
        raise Exception(input1_sheet.cell(row=tmpRow, column=1).value)

    tmpRow =  tmpRow + 1

rows = input2_sheet.max_row
tmpRow = 2
pack2 = dict()
while tmpRow <= rows:
    if input2_sheet.cell(row=tmpRow, column=1).value != None:
        #过滤 =A/B 字符
        key = myfilter(input2_sheet.cell(row=tmpRow, column=1).value,input2_sheet)

        name = (input2_sheet.cell(row=tmpRow, column=2).value)
        if key not in pack2.keys():
            pack2[key] = (
                myfilter(name,input2_sheet),
                int(input2_sheet.cell(row=tmpRow, column=3).value) +
                int(input2_sheet.cell(row=tmpRow, column=4).value) -
                int(input2_sheet.cell(row=tmpRow, column=5).value)
            )
        else:
            pack2[key] = (
                pack2[key][0],
                pack2[key][1] +
                int(input2_sheet.cell(row=tmpRow, column=3).value) +
                int(input2_sheet.cell(row=tmpRow, column=4).value) -
                int(input2_sheet.cell(row=tmpRow, column=5).value)
            )
    else:
        raise Exception(input2_sheet.cell(row=tmpRow, column=1).value)

    tmpRow =  tmpRow + 1


#入库单明细表
# print(pack1)

#进销存表
# print(pack2)



#############

#入库单明细表存在， 而未在进销存表
pack1NotInPack2 = None

#进销存表存在，而入库单明细表为存在
pack2NotInPack1 = None

#所有同时在进销存表 和 入库单明细表 的 商品
pack1SamePack2 = None

#####################################
#所有pack1 不在 pack2 里面的
pack1NotInPack2 =  dict()
for key1 in pack1:
       if key1 not in pack2.keys():
           pack1NotInPack2[key1] = pack1[key1]


#所有pack2 不在 pack1 里面的
pack2NotInPack1 =  dict()
for key2 in pack2:
       if key2 not in pack1.keys():
           pack2NotInPack1[key2] = pack2[key2]


# print(pack1NotInPack2)
# print(pack2NotInPack1)
#######################################


######################################
# 所有pack1 和 pack2 相同的商品条码
pack1SamePack2 =  dict()
for key in pack1:
       if key in pack2.keys():
           pack1SamePack2[key] = (pack1[key][0],int(pack1[key][1]), int(pack2[key][1]), int(pack2[key][1]) - int(pack1[key][1]) )

# print(pack1SamePack2)
#######################################

####test###
# for key in pack1SamePack2:
#     if key in pack1NotInPack2.keys():
#         print("zx")
#     if key in pack2NotInPack1.keys():
#         print("qq")


# 默认表sheet1
newWb = Workbook()
ws1 = newWb.active
ws1.cell(row=1, column=1, value='商品条码')
ws1.cell(row=1, column=2, value='产品名称')
ws1.cell(row=1, column=3, value='佳成数量')
ws1.cell(row=1, column=4, value='御商数量')
ws1.cell(row=1, column=5, value='偏差')

#填写
tmpRow = 2
for key in pack1SamePack2:
    ws1.cell(row=tmpRow, column=1, value=key)
    ws1.cell(row=tmpRow, column=2, value=pack1SamePack2[key][0])
    ws1.cell(row=tmpRow, column=3, value=pack1SamePack2[key][1])
    ws1.cell(row=tmpRow, column=4, value=pack1SamePack2[key][2])
    ws1.cell(row=tmpRow, column=5, value=pack1SamePack2[key][3])

    tmpRow = tmpRow + 1

##########
for key in pack2NotInPack1:
    ws1.cell(row=tmpRow, column=1, value=key)
    ws1.cell(row=tmpRow, column=2, value=pack2NotInPack1[key][0])
    ws1.cell(row=tmpRow, column=4, value=pack2NotInPack1[key][1])
    ws1.cell(row=tmpRow, column=3, value="-")
    ws1.cell(row=tmpRow, column=5, value=("error"))

    tmpRow = tmpRow + 1


##########
for key in pack1NotInPack2:
    ws1.cell(row=tmpRow, column=1, value=key)
    ws1.cell(row=tmpRow, column=2, value=pack1NotInPack2[key][0])
    ws1.cell(row=tmpRow, column=3, value=pack1NotInPack2[key][1])
    ws1.cell(row=tmpRow, column=4, value="-")
    ws1.cell(row=tmpRow, column=5, value=("error"))

    tmpRow = tmpRow + 1

############
newWb.save("C:/Users/admin/Desktop/output.xlsx")